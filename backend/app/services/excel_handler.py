"""
Excel import/export service for bulk question uploads and result downloads.
"""

from datetime import datetime
from io import BytesIO
from sqlalchemy.orm import Session
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from ..models.quiz import Quiz
from ..models.question import Question
from ..models.answer import Answer
from ..models.quiz_attempt import QuizAttempt
from ..models.student import Student


def import_excel_questions(
    db: Session,
    file_content: bytes,
    quiz_name: str,
    quiz_description: str = None,
    category_id: int = None,
):
    """
    Import questions from Excel file.
    Expected format: Savol | A | B | C | D | To'g'ri
    
    Args:
        db: Database session
        file_content: .xlsx file bytes
        quiz_name: Name for the quiz
        quiz_description: Optional description
        category_id: Optional category ID
    
    Returns: Created quiz object with questions
    
    Raises:
        ValueError: If file format is invalid
    """
    
    try:
        # Load workbook
        workbook = load_workbook(BytesIO(file_content))
        
        # Get the "Questions" sheet or first sheet
        sheet = None
        for ws in workbook.sheetnames:
            if ws.lower() == "questions":
                sheet = workbook[ws]
                break
        
        if not sheet:
            sheet = workbook.active
        
        if not sheet:
            raise ValueError("No worksheet found in Excel file")
        
        # Parse rows
        questions_data = []
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), 2):
            if not row or all(v is None for v in row):
                continue
            
            if len(row) < 6:
                raise ValueError(
                    f"Row {row_idx}: Expected 6 columns (Savol, A, B, C, D, To'g'ri), "
                    f"got {len(row)}"
                )
            
            question_text, opt_a, opt_b, opt_c, opt_d, correct = row[:6]
            
            if not question_text:
                raise ValueError(f"Row {row_idx}: Question text is required")
            
            if not all([opt_a, opt_b, opt_c, opt_d]):
                raise ValueError(
                    f"Row {row_idx}: All 4 options (A, B, C, D) are required"
                )
            
            correct_upper = str(correct).upper().strip() if correct else None
            if correct_upper not in ["A", "B", "C", "D"]:
                raise ValueError(
                    f"Row {row_idx}: Correct answer must be A, B, C, or D, got '{correct}'"
                )
            
            questions_data.append({
                "question": str(question_text).strip(),
                "options": [
                    str(opt_a).strip(),
                    str(opt_b).strip(),
                    str(opt_c).strip(),
                    str(opt_d).strip(),
                ],
                "correct_index": ord(correct_upper) - ord("A"),  # 0-3
            })
        
        if not questions_data:
            raise ValueError("No questions found in Excel file")
        
        # Create quiz
        quiz = Quiz(
            title=quiz_name,
            description=quiz_description or "",
            category=category_id,
            status="published",
            shuffle_questions=False,
            shuffle_answers=False,
            show_result=True,
        )
        db.add(quiz)
        db.flush()  # Get quiz.id without commit
        
        # Create questions and answers
        for order, q_data in enumerate(questions_data, 1):
            question = Question(
                quiz_id=quiz.id,
                text=q_data["question"],
                points=1,
                order_number=order,
            )
            db.add(question)
            db.flush()
            
            # Create answers
            for option_idx, option_text in enumerate(q_data["options"]):
                is_correct = option_idx == q_data["correct_index"]
                answer = Answer(
                    question_id=question.id,
                    text=option_text,
                    is_correct=is_correct,
                )
                db.add(answer)
        
        db.commit()
        
        return {
            "success": True,
            "quiz_id": quiz.id,
            "quiz_name": quiz.title,
            "questions_created": len(questions_data),
            "status": quiz.status,
        }
    
    except Exception as e:
        db.rollback()
        raise ValueError(f"Excel import error: {str(e)}")


def export_attempts_to_excel(
    db: Session,
    start_date: datetime = None,
    end_date: datetime = None,
    quiz_id: int = None,
    category_id: int = None,
    student_id: int = None,
):
    """
    Export quiz attempts to Excel file.
    
    Args:
        db: Database session
        start_date: Filter from this date
        end_date: Filter until this date
        quiz_id: Filter by specific quiz
        category_id: Filter by specific category
        student_id: Filter by specific student
    
    Returns: BytesIO object containing .xlsx file
    """
    
    # Build query
    query = db.query(
        Student.name,
        Quiz.title,
        QuizAttempt.score,
        QuizAttempt.percentage,
        QuizAttempt.duration,
        QuizAttempt.finished_at,
    ).join(Student).join(Quiz).filter(QuizAttempt.finished_at.isnot(None))
    
    # Apply filters
    if start_date:
        query = query.filter(QuizAttempt.finished_at >= start_date)
    
    if end_date:
        query = query.filter(QuizAttempt.finished_at <= end_date)
    
    if quiz_id:
        query = query.filter(QuizAttempt.quiz_id == quiz_id)
    
    if category_id:
        query = query.filter(Quiz.category == category_id)
    
    if student_id:
        query = query.filter(QuizAttempt.student_id == student_id)
    
    attempts = query.order_by(QuizAttempt.finished_at.desc()).all()
    
    # Create workbook
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Attempts"
    
    # Add header row
    headers = [
        "Student Name",
        "Quiz Name",
        "Score",
        "Percentage",
        "Duration (min)",
        "Date",
    ]
    
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col_idx, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col_idx)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    # Add data rows
    for row_idx, attempt in enumerate(attempts, 2):
        student_name, quiz_name, score, percentage, duration, finished_at = attempt
        
        duration_minutes = duration // 60 if duration else 0
        date_str = finished_at.strftime("%Y-%m-%d %H:%M") if finished_at else ""
        
        sheet.cell(row=row_idx, column=1).value = student_name
        sheet.cell(row=row_idx, column=2).value = quiz_name
        sheet.cell(row=row_idx, column=3).value = score
        sheet.cell(row=row_idx, column=4).value = f"{percentage:.1f}%" if percentage else "0%"
        sheet.cell(row=row_idx, column=5).value = duration_minutes
        sheet.cell(row=row_idx, column=6).value = date_str
        
        # Center align numeric columns
        for col in [3, 4, 5]:
            sheet.cell(row=row_idx, column=col).alignment = Alignment(horizontal="center")
    
    # Adjust column widths
    sheet.column_dimensions["A"].width = 20
    sheet.column_dimensions["B"].width = 25
    sheet.column_dimensions["C"].width = 12
    sheet.column_dimensions["D"].width = 12
    sheet.column_dimensions["E"].width = 15
    sheet.column_dimensions["F"].width = 20
    
    # Save to BytesIO
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    
    return output


def export_question_template_to_excel():
    """
    Export a blank template for bulk question import.
    
    Returns: BytesIO object containing template .xlsx
    """
    
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Questions"
    
    # Add headers
    headers = ["Savol (Question)", "A (Option 1)", "B (Option 2)", "C (Option 3)", "D (Option 4)", "To'g'ri (Correct)"]
    
    header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col_idx, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col_idx)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    # Add example row
    example = [
        "Capital of France?",
        "London",
        "Paris",
        "Berlin",
        "Madrid",
        "B",
    ]
    
    for col_idx, value in enumerate(example, 1):
        cell = sheet.cell(row=2, column=col_idx)
        cell.value = value
        cell.font = Font(italic=True, color="999999")
    
    # Add empty rows for data entry
    for row in range(3, 12):
        for col in range(1, 7):
            sheet.cell(row=row, column=col)
    
    # Adjust column widths
    sheet.column_dimensions["A"].width = 25
    sheet.column_dimensions["B"].width = 20
    sheet.column_dimensions["C"].width = 20
    sheet.column_dimensions["D"].width = 20
    sheet.column_dimensions["E"].width = 20
    sheet.column_dimensions["F"].width = 15
    
    # Save to BytesIO
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    
    return output
